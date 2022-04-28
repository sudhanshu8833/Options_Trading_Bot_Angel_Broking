import time,datetime
from kiteconnect import KiteConnect,KiteTicker
from selenium import webdriver
from selenium.webdriver.common.by import By

import openpyxl
import pathlib
# request_token='weBVhRnaSQadQA6J9QC5LVBWWQcgp8Nt'
apiKey='nu3ion96bs510ej9'
secretApi='1dp9cdmj7a8sr4pji0yy9oy6lkr00uun'
userId='WT1413'
userPass='8826411011'
pin='960220'
# kite=kiteconnect.KiteConnect(api_key=api_key)
# data=kite.generate_session(request_token, api_secret=secret_api_key)



def getAccessToken(userId,userPass,pin,apiKey,secretApi):
    kite=KiteConnect(api_key=apiKey)
    url=kite.login_url()
    print(url)
    try:
        
        # browser=webdriver.Firefox(executable_path=r"E:\Python\jpdelhi\exceltokite\geckodriver.exe")
        browser=webdriver.Safari(executable_path='/Users/sudhanshu/Desktop/algo-trading/PROJECTS/PROJECT4/chromedriver')
        browser.get(url)
        # time.sleep(3)
        userIdInput = browser.find_element(By.XPATH, '//*[@id="container"]/div/div/div[2]/form/div[1]/input')
        userIdInput.send_keys(userId)
        passInput = browser.find_element(By.XPATH, '//*[@id="container"]/div/div/div[2]/form/div[2]/input')
        passInput.send_keys(userPass)
        nextBtn = browser.find_element(By.XPATH, '//*[@id="container"]/div/div/div[2]/form/div[4]/button')
        nextBtn.click()
        time.sleep(7)
        twofapin = browser.find_element(By.XPATH, '//*[@id="container"]/div/div/div[2]/form/div[2]/div/input')
        twofapin.send_keys(pin)
        submitBtn = browser.find_element(By.XPATH, '//*[@id="container"]/div/div/div[2]/form/div[3]/button')
        submitBtn.click()
        time.sleep(10)
        print(browser.current_url)
        token_url = browser.current_url
        browser.quit()
        print(token_url)
        if token_url.find('success'):
            l=token_url.find('request_token')
            print(l)
            if l!=-1:
                print(l)
                nextAmpersand=token_url.find('&',l)
                nextAmpersand=None if nextAmpersand==-1 else nextAmpersand
                requestToken=token_url[l+14:nextAmpersand]
                try:
                    response=kite.generate_session(requestToken, api_secret=secretApi)
                    return {'status': 1,'accessToken':response['access_token']}
                except Exception as e:
                    return {'status':0,'msg':f'{e} - Error While Fetching Access Token'}
    except Exception as e:
        print(e)
        {'status':0,'msg':f'{e} Error'}
    return {'status': 0,'msg':'Unable to Fetch Request_token'}


def login(userId,userPass,pin,apiKey,secretApi,accessToken=None,tickerFlag=0):
    try:
        print('hi')
        print(userPass)
        ticker=None
        kite=KiteConnect(api_key=apiKey,access_token=accessToken)
        kite.orders()
        if tickerFlag:
            ticker=KiteTicker(api_key=apiKey,access_token=accessToken)
        return {'status':1,'kite':kite,'ticker':ticker}
    except Exception as e:
        print('error login')
        accessTokenResult=getAccessToken(userId,userPass,pin,apiKey,secretApi)
        print(accessTokenResult)
        if accessTokenResult.get('status'):
            kite=KiteConnect(api_key=apiKey,access_token=accessTokenResult.get('accessToken'))
            if tickerFlag:
                ticker=KiteTicker(api_key=apiKey,access_token=accessTokenResult.get('accessToken'))
            return {'status':2,'kite':kite,'accessToken':accessTokenResult.get('accessToken'),'ticker':ticker}
        return {'status':0,'msg':accessTokenResult.get('msg')}


def loginThroughFile(zerodhaUserId,wb,tickerFlag=0):
    pointer=2
    # wb=xlwings.Book(r"E:\Python\jpdelhi\exceltokite\Trade Setup File.xlsx")
    ws=wb.sheets['Zerodha Users']
    
    while ws.range(f'A{pointer}').value!=None:
        if ws.range(f'A{pointer}').value==zerodhaUserId:
            accessTokenTimeStamp=ws.range(f'G{pointer}').value
            # if accessTokenTimeStamp==datetime.datetime.today():
            accessToken=ws.range(f'F{pointer}').value

            userId=ws.range(f'A{pointer}').value
            userPass=str(ws.range(f'B{pointer}').value)
            pin=str(ws.range(f'C{pointer}').value)
            apiKey=ws.range(f'D{pointer}').value
            apiSecret=ws.range(f'E{pointer}').value
            loginResult=login(userId,userPass,pin,apiKey,apiSecret,accessToken,tickerFlag=tickerFlag)
            print(loginResult)
            if loginResult.get('status')==2:
                ws.range(f'F{pointer}').value=loginResult.get('accessToken')
                ws.range(f'G{pointer}').value=datetime.datetime.today()
                wb.save()
            return loginResult
        else:
            pointer+=1
    return {'status':0,'msg':'Zerodha Userid Not Found in File'}


def loginEasy(zerodhaUserId,file,tickerFlag=0):
    wb1=openpyxl.load_workbook(file,data_only=True)
    ws=wb1['Zerodha Users']
    print(file)
    pointer=2
    while ws.cell(row=pointer,column=1).value!=None:
        if ws.cell(row=pointer,column=1).value==zerodhaUserId:
            userId=ws.cell(row=pointer,column=1).value
            userPass=ws.cell(row=pointer,column=2).value
            pin=ws.cell(row=pointer,column=3).value
            apiKey=ws.cell(row=pointer,column=4).value
            apiSecret=ws.cell(row=pointer,column=5).value
            accessToken=ws.cell(row=pointer,column=6).value
            loginResult=login(userId,userPass,pin,apiKey,apiSecret,accessToken,tickerFlag=tickerFlag)
            print(loginResult)
            if loginResult.get('status')==2: #Login Via Browser
                ws.cell(row=pointer,column=6).value=loginResult.get('accessToken')
                ws.cell(row=pointer,column=7).value=datetime.datetime.now()
                try:
                    
                    wb1.save(file)
                    print('Saved')
                except Exception as e:
                    print(e)
                    loginResult['msg']='Users File is Open Unable to save it.'
                    
            return loginResult
        else:
            pointer+=1
    return {'status':0,'msg':'Zerodha User Id Not Found in given File'}
    




if __name__ == '__main__':
    login(userId,userPass,pin,apiKey,secretApi)
    # print(loginEasy('NO2904',r"G:\Python\projects\jpdelhi\optionWithoutHedging\Users.xlsx"))










# print(data)
# print(z)

# z=hashlib.sha256(api_key.encode('utf-8') + request_token.encode('utf-8') + secret_api_key.encode('utf-8'))
# checksum=z.hexdigest()
# k={
#     'api_key': api_key,
#     'request_token': request_token,
#     'checksum': checksum
# }
# print(params)
# r=requests.post('https://api.kite.trade/session/token',params=k)
# print(r.content)
# print(r.get(url).headers)
